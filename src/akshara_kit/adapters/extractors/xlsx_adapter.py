"""XLSX extraction via openpyxl.

One deterministic path — there is nothing to race — so this adapter is also
the XLSX coordinator. It handles font detection and legacy normalisation
inline, because openpyxl gives a font per cell, which is exactly the
granularity conversion needs.
"""

from __future__ import annotations

import time
from typing import TYPE_CHECKING, Iterator

from akshara_kit.contracts.extraction import ExtractionResult, SourceFormat
from akshara_kit.eye.errors import AdapterUnavailableError, ExtractionFailedError
from akshara_kit.eye.quality_probe import score

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook

    from akshara_kit.eye.font_detection import SpanFont

__all__ = ["BACKEND_ID", "extract"]

BACKEND_ID = "openpyxl"

#: Separators. Cells within a row are tab-joined so a row stays one line.
_CELL_SEPARATOR = "\t"
_ROW_SEPARATOR = "\n"
_SHEET_SEPARATOR = "\n\n"


def extract(file_path: str) -> ExtractionResult:
    """Extract every non-empty cell across every worksheet.

    Realises Section 5's XLSX path and Section 6.3's per-cell font detection.
    """
    from akshara_kit.eye.encoding_normaliser import normalise_spans

    started = time.perf_counter()
    workbook = _open(file_path)
    try:
        spans = list(_iter_cell_spans(workbook))
        sheet_names = [sheet.title for sheet in workbook.worksheets]
    finally:
        workbook.close()

    outcome = normalise_spans(spans)
    return ExtractionResult(
        text=outcome.text,
        backend_id=BACKEND_ID,
        source_format=SourceFormat.XLSX,
        latency_seconds=time.perf_counter() - started,
        quality=score(outcome.text),
        font_detection_method=outcome.method,
        detected_legacy_fonts=outcome.converted_fonts,
        unmapped_legacy_fonts=outcome.unmapped_fonts,
        metadata={"sheets": sheet_names},
    )


def _open(file_path: str) -> Workbook:
    """Load the workbook, translating library failures into named errors."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on install extras
        raise AdapterUnavailableError(
            "openpyxl is not installed; install the 'xlsx' extra"
        ) from exc

    try:
        # data_only=True yields cached formula results rather than "=SUM(A1:A9)".
        return load_workbook(file_path, data_only=True)
    except Exception as exc:
        raise ExtractionFailedError(
            f"{BACKEND_ID} could not open {file_path}: {exc}"
        ) from exc


def _iter_cell_spans(workbook: Workbook) -> Iterator[SpanFont]:
    """Yield one span per non-empty cell, plus layout separator spans.

    Worksheet order is openpyxl's, which preserves the workbook's own order.
    """
    from akshara_kit.eye.font_detection import SpanFont, separator

    for sheet_index, sheet in enumerate(workbook.worksheets):
        if sheet_index:
            yield separator(_SHEET_SEPARATOR)
        yield separator(f"=== {sheet.title} ==={_ROW_SEPARATOR}")

        for row in sheet.iter_rows():
            populated = [cell for cell in row if cell.value is not None]
            if not populated:
                continue
            for cell_index, cell in enumerate(populated):
                if cell_index:
                    yield separator(_CELL_SEPARATOR)
                yield SpanFont(
                    text=str(cell.value),
                    font=_cell_font(cell),
                    location=f"{sheet.title}!{cell.coordinate}",
                )
            yield separator(_ROW_SEPARATOR)


def _cell_font(cell) -> str:
    """The cell's font name, or ``""`` when it inherits the workbook default."""
    font = getattr(cell, "font", None)
    return getattr(font, "name", None) or ""
